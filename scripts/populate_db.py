# NOME COMPLETO B | SEXO E | DATA NASC F | TELEFONE G | EMAIL I | CULTO K | MEMBRO L | SIGI Q | NOME EM FOTO S

from PIL import Image
import face_recognition
import numpy as np
import openpyxl
import argparse
import os
import io
import re

from database.Mongodb import Mongodb

from entities.Person import Person
from entities.Gender import Gender
from entities.Calendar import Calendar
from entities.Day import Day
from entities.Collections import Collections
from entities.Photo import Photo
from entities.PhotoCategory import PhotoCategory
from entities.PhotoMode import PhotoMode
from entities.Name import Name
from entities.Ministry import Ministry
from entities.Encoding import Encoding

class Workbook:
    rows = range(2, 200)
    columns = 'BEFGIKLQS'

    def __init__(self, fp, sheet):
        self.wb = openpyxl.load_workbook(fp)
        self.ws = self.wb.get_sheet_by_name(sheet)

    def list_of_people(self):
        l = []
        for row in Workbook.rows:
            p = []
            for column in Workbook.columns:
                cell = column + str(row)
                value = self.ws[cell].value
                p.append(value)
            if p[0] is not None and p[-1] is not None:
                l.append(p)
        return l

    def dict_of_people(self, l):
        d = []

        for element in l:
            name = Name.from_str(element[0])

            gender = Gender.FEMALE if element[1] == 'F' else Gender.MALE

            day = element[2]
            if day is None:
                birth_date = None
            else:
                birth_date = Day(day.year, day.month, day.day)

            if element[3] is None:
                phone_number = None
            elif type(element[3]) == float:
                phone_number = str(element[3])[:-2]
            elif type(element[3]) == int:
                phone_number = str(element[3])
            else:
                phone_number = re.compile('[\W_]+').sub('', element[3])

            email = element[4]

            ministry = [Ministry[x.strip().upper()] for x in element[5].split(',')]

            member = bool(element[6])

            sigi = element[7]

            calendar = Calendar()

            photos = {
                PhotoCategory.FRONT: Photo(PhotoCategory.FRONT, PhotoMode.RAW, 0, element[8].lower() + '-fr.jpg'),
                PhotoCategory.LEFT: Photo(PhotoCategory.LEFT, PhotoMode.RAW, 0, element[8].lower() + '-le.jpg'),
                PhotoCategory.RIGHT: Photo(PhotoCategory.RIGHT, PhotoMode.RAW, 0, element[8].lower() + '-ld.jpg'),
            }

            encodings = {
                PhotoCategory.FRONT: Photo(PhotoCategory.FRONT, PhotoMode.RAW, 0, element[8].lower() + '-fr.jpg'),
                PhotoCategory.LEFT: Photo(PhotoCategory.LEFT, PhotoMode.RAW, 0, element[8].lower() + '-le.jpg'),
                PhotoCategory.RIGHT: Photo(
                    PhotoCategory.RIGHT, PhotoMode.RAW, 0, element[8].lower() + '-ld.jpg')
            }

            d.append(Person(name, birth_date, email, gender, phone_number,
                            member, ministry, sigi, calendar, photos, encodings))

        return d

class Rotate:
    @staticmethod
    def rotate(image):
        image_exif = image._getexif()
        image_orientation = image_exif[274]
        if image_orientation == 2:
            image = image.transpose(Image.FLIP_LEFT_RIGHT)
        if image_orientation == 3:
            image = image.transpose(Image.ROTATE_180)
        if image_orientation == 4:
            image = image.transpose(Image.FLIP_TOP_BOTTOM)
        if image_orientation == 5:
            image = image.transpose(
                Image.FLIP_LEFT_RIGHT).transpose(Image.ROTATE_90)
        if image_orientation == 6:
            image = image.transpose(Image.ROTATE_270)
        if image_orientation == 7:
            image = image.transpose(
                Image.FLIP_TOP_BOTTOM).transpose(Image.ROTATE_90)
        if image_orientation == 8:
            image = image.transpose(Image.ROTATE_90)
        return image

def rename_lower(path):
    for f in os.listdir(path):
        os.rename(os.path.join(path, f), os.path.join(path, f.lower()))

def populate(d, db, args):
    for person in d:
        print('saving person...', end=' ')
        member_id = db.insert_member(person)

        encoding_saved = False
        print(person.name)

        for key in person.photos.keys():
            if person.photos[key] is None:
                continue
            else:
                try:
                    fpath = os.path.join(args.path, person.photos[key].data.lower())

                    foto = Image.open(fpath)
                    try:
                        foto = Rotate.rotate(foto)
                    except Exception as e:
                        pass
                    encodings = face_recognition.face_encodings(np.array(foto))[0]

                    print('saving %s: %s...' % (key, fpath), end=' ')
                    encoding = Encoding(member_id, person.name, encodings)

                    encoding_id = db.insert_encoding(encoding)
                    person.photos[key] = encoding_id

                    imgByteArr = io.BytesIO()
                    foto.save(imgByteArr, format='JPEG')
                    image_id = db.insert_image(imgByteArr.getvalue())
                    person.encodings[key] = image_id

                    encoding_saved = True
                    print('done')
                except Exception as e:
                    print(e)
                    continue

        print('updating person...')
        db.replace_member(member_id, person)

def createArgsParser():
    parser = argparse.ArgumentParser()
    parser.add_argument('file', help="xlsx file")
    parser.add_argument('-s', '--sheet', required=True, help='xlsx sheet')
    parser.add_argument('-p', '--path', required=True,
                        help='folder with pictures')
    parser.add_argument('-d', '--database', required=True,
                        help='database name')
    return parser.parse_args()

if __name__ == '__main__':
    args = createArgsParser()
    rename_lower(args.path)

    wb = Workbook(args.file, args.sheet)
    d = wb.dict_of_people(wb.list_of_people())

    db = Mongodb(db=args.database)
    db.delete_all_encodings(True)
    db.delete_all_members(True)

    populate(d, db, args)
